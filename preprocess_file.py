# excel and table detection code

import re
import logging
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple
import sys
import json
import pandas as pd
import numpy as np
from reportlab.lib import colors
from openpyxl import load_workbook
from openpyxl.styles.fills import GradientFill, PatternFill
from openpyxl.worksheet.worksheet import Worksheet
from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
from reportlab.lib.units import cm
from reportlab.lib.enums import TA_CENTER, TA_LEFT
from reportlab.lib.pagesizes import A4
from reportlab.platypus import (
    HRFlowable,
    Image,
    PageBreak,
    Paragraph,
    SimpleDocTemplate,
    Spacer,
    Table,
    TableStyle,
)

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [%(levelname)s] %(message)s",
    handlers=[logging.StreamHandler(sys.stdout)],
)
log = logging.getLogger(__name__)


def detect_numeric_columns(df: pd.DataFrame, sample_size: int = 3) -> List[str]:
    """
    Detects columns that can be treated as numeric.

    The function checks a small sample of non-empty values from each column,
    removes common formatting symbols such as commas, currency symbols, and
    percentages, then tests whether most sampled values can be converted to float.

    Args:
        df (pd.DataFrame): DataFrame whose columns need to be checked.
        sample_size (int): Number of non-empty values to sample from each column.

    Returns:
        List[str]: Column names that appear to contain numeric values.
    """
    numeric_cols = []

    for col in df.columns:
        if col.startswith("__"):
            continue
        sample = (
            df[col]
            .dropna()
            .astype(str)
            .str.strip()
        )

        sample = sample[sample != ""].head(sample_size)

        if len(sample) == 0:
            continue

        success = 0

        for val in sample:
            cleaned = (
                val.replace(",", "")
                   .replace("$", "")
                   .replace("₹", "")
                   .replace("%", "")
            )

            try:
                float(cleaned)
                success += 1
            except:
                pass
        if success / len(sample) >= 0.7:
            numeric_cols.append(col)

    return numeric_cols


def _analyze_cell_formatting(sheet: Worksheet, row: int, col: int) -> Dict[str, Any]:
    """
    Reads formatting signals from a single Excel cell.

    This helper checks whether a cell has borders, bold text, fill color,
    merged-cell status, alignment, font size, and value type. These signals are
    later used to identify formatted table headers and table regions.

    Args:
        sheet (Worksheet): OpenPyXL worksheet to inspect.
        row (int): One-based row number of the target cell.
        col (int): One-based column number of the target cell.

    Returns:
        Dict[str, Any]: Formatting metadata for the selected cell.
    """
    cell = sheet.cell(row=row, column=col)
    fmt: Dict[str, Any] = {
        "has_border": False,
        "is_bold": False,
        "has_fill": False,
        "is_merged": False,
        "alignment": None,
        "font_size": None,
        "value_type": type(cell.value).__name__ if cell.value is not None else "NoneType",
    }
    if cell.border and any([
        getattr(cell.border.left, "style", None),
        getattr(cell.border.right, "style", None),
        getattr(cell.border.top, "style", None),
        getattr(cell.border.bottom, "style", None),
    ]):
        fmt["has_border"] = True
    if cell.font:
        fmt["is_bold"] = cell.font.bold or False
        fmt["font_size"] = cell.font.size
    if cell.fill:
        if isinstance(cell.fill, PatternFill):
            if cell.fill.start_color and cell.fill.start_color.index != "00000000":
                fmt["has_fill"] = True
        elif isinstance(cell.fill, GradientFill):
            if cell.fill.stop and any(str(c) != "00000000" for c in cell.fill.stop):
                fmt["has_fill"] = True
    for merged_range in sheet.merged_cells.ranges:
        if cell.coordinate in merged_range:
            fmt["is_merged"] = True
            break
    if cell.alignment:
        fmt["alignment"] = cell.alignment.horizontal
    return fmt


def _build_merge_map(sheet_format: Worksheet, sheet_values: Worksheet) -> Dict[Tuple[int, int], Any]:
    """
    Builds a lookup map for merged-cell values.

    For every merged range in the formatted worksheet, this function maps each
    cell inside the merged range to the top-left value from the values worksheet.
    This helps preserve header or data values that visually span multiple cells.

    Args:
        sheet_format (Worksheet): Worksheet loaded with formulas and formatting.
        sheet_values (Worksheet): Worksheet loaded with calculated values.

    Returns:
        Dict[Tuple[int, int], Any]: Mapping from cell coordinates to merged values.
    """
    merge_map: Dict[Tuple[int, int], Any] = {}
    for merged_range in sheet_format.merged_cells.ranges:
        top_left_val = sheet_values.cell(
            row=merged_range.min_row, column=merged_range.min_col
        ).value
        for r in range(merged_range.min_row, merged_range.max_row + 1):
            for c in range(merged_range.min_col, merged_range.max_col + 1):
                merge_map[(r, c)] = top_left_val
    return merge_map


def _find_formatted_table_boundary(
    sheet: Worksheet,
    format_grid: Dict,
    start_row: int,
    start_col: int,
    processed_cells: set,
) -> Optional[Tuple[int, int, int, int]]:
    """
    Finds the rectangular boundary of a formatted table.

    Starting from a likely header cell, this function expands left and right to
    locate the header width, then scans downward until the data region ends.

    Args:
        sheet (Worksheet): Worksheet being analyzed.
        format_grid (Dict): Formatting metadata for non-empty cells.
        start_row (int): Candidate header row.
        start_col (int): Candidate header column.
        processed_cells (set): Cells already assigned to previously detected tables.

    Returns:
        Optional[Tuple[int, int, int, int]]: Table region as
        (start_row, end_row, start_col, end_col), or None if no valid table is found.
    """
    max_row = sheet.max_row
    max_col = sheet.max_column

    header_start_col = start_col
    header_end_col = start_col

    for col in range(start_col - 1, 0, -1):
        if (start_row, col) in format_grid and sheet.cell(row=start_row, column=col).value is not None:
            header_start_col = col
        else:
            break

    for col in range(start_col + 1, max_col + 1):
        if (start_row, col) in format_grid and sheet.cell(row=start_row, column=col).value is not None:
            header_end_col = col
        else:
            break

    data_end_row = start_row
    for row in range(start_row + 1, max_row + 1):
        has_data = any(
            sheet.cell(row=row, column=col).value is not None
            for col in range(header_start_col, header_end_col + 1)
        )
        if has_data:
            data_end_row = row
        else:
            empty_rows = sum(
                1 for nr in range(row, min(row + 3, max_row + 1))
                if not any(
                    sheet.cell(row=nr, column=col).value is not None
                    for col in range(header_start_col, header_end_col + 1)
                )
            )
            if empty_rows >= 2:
                break

    if data_end_row > start_row and header_end_col > header_start_col:
        return (start_row, data_end_row, header_start_col, header_end_col)
    return None


def _detect_formatted_table_regions(
    sheet: Worksheet,
    min_rows: int = 3,
    min_cols: int = 3,
) -> List[Tuple[int, int, int, int]]:
    """
    Detects table-like regions using Excel formatting cues.

    The function scans non-empty cells, identifies possible header cells based on
    formatting such as borders, bold text, fills, or merged cells, and then expands
    each candidate into a rectangular table boundary.

    Args:
        sheet (Worksheet): Worksheet to inspect.
        min_rows (int): Minimum number of rows required for a valid table.
        min_cols (int): Minimum number of columns required for a valid table.

    Returns:
        List[Tuple[int, int, int, int]]: Detected table regions as
        (start_row, end_row, start_col, end_col).
    """
    max_row = sheet.max_row
    max_col = sheet.max_column
    if max_row == 1 and max_col == 1:
        return []

    format_grid: Dict[Tuple[int, int], Dict] = {}
    for row in range(1, max_row + 1):
        for col in range(1, max_col + 1):
            cell = sheet.cell(row=row, column=col)
            if cell.value is not None:
                format_grid[(row, col)] = _analyze_cell_formatting(sheet, row, col)

    header_candidates = [
        (row, col)
        for (row, col), fmt in format_grid.items()
        if (fmt["has_border"] or fmt["is_bold"] or fmt["has_fill"] or fmt["is_merged"])
        and sheet.cell(row=row, column=col).value
    ]

    table_regions: List[Tuple[int, int, int, int]] = []
    processed_cells: set = set()

    for row, col in header_candidates:
        if (row, col) in processed_cells:
            continue
        region = _find_formatted_table_boundary(sheet, format_grid, row, col, processed_cells)
        if region:
            sr, er, sc, ec = region
            if (er - sr + 1) >= min_rows and (ec - sc + 1) >= min_cols:
                table_regions.append(region)
                for r in range(sr, er + 1):
                    for c in range(sc, ec + 1):
                        processed_cells.add((r, c))

    return table_regions


def _col_letter(col_idx: int) -> str:
    """
    Converts a zero-based column index into an Excel column letter.

    Args:
        col_idx (int): Zero-based column index.

    Returns:
        str: Excel-style column label, such as A, B, Z, AA, or AB.
    """
    if col_idx < 0:
        col_idx = 0
    result = ""
    col_idx += 1
    while col_idx:
        col_idx, remainder = divmod(col_idx - 1, 26)
        result = chr(65 + remainder) + result
    return result


def _detect_tables_in_sheet_values(
    sheet_data: List[List],
    min_rows: int = 3,
    min_cols: int = 2,
) -> List[str]:
    """
    Detects table ranges from raw worksheet values.

    This fallback method identifies consecutive row blocks where each row has at
    least a minimum number of non-empty cells, then converts those blocks into
    Excel-style ranges.

    Args:
        sheet_data (List[List]): Raw worksheet values as a list of rows.
        min_rows (int): Minimum number of rows required for a detected table.
        min_cols (int): Minimum number of non-empty cells required per row.

    Returns:
        List[str]: Excel-style table ranges such as "A1:D20".
    """
    df = pd.DataFrame(sheet_data)
    non_empty = df.notna() & df.map(lambda x: str(x).strip() != "")

    table_ranges: List[str] = []
    in_table = False
    start_row = None

    def _add_range(s_row: int, e_row: int) -> None:
        """
        Adds a detected row block as an Excel-style range.

        Args:
            s_row (int): Zero-based starting row index.
            e_row (int): Zero-based ending row index.

        Returns:
            None
        """
        tdf = df.iloc[s_row: e_row + 1]
        has_data = tdf.apply(lambda c: c.notna().sum() > 0).values
        if not has_data.any():
            return
        sc = next(i for i, v in enumerate(has_data) if v)
        ec = len(has_data) - next(i for i, v in enumerate(reversed(has_data)) if v) - 1
        if e_row - s_row + 1 >= min_rows:
            table_ranges.append(
                f"{_col_letter(sc)}{s_row + 1}:{_col_letter(ec)}{e_row + 1}"
            )

    for i, row in non_empty.iterrows():
        if row.sum() >= min_cols:
            if not in_table:
                in_table = True
                start_row = i
        else:
            if in_table:
                _add_range(start_row, i - 1)
                in_table = False

    if in_table:
        _add_range(start_row, len(df) - 1)

    return table_ranges


def _extract_formatted_table(
    sheet_values: Worksheet,
    sheet_format: Worksheet,
    region: Tuple[int, int, int, int],
) -> pd.DataFrame:
    """
    Extracts a formatted Excel table region into a DataFrame.

    The first row of the region is treated as the header row. Merged-cell values
    and formula text are preserved where normal calculated values are missing.

    Args:
        sheet_values (Worksheet): Worksheet loaded with calculated cell values.
        sheet_format (Worksheet): Worksheet loaded with formulas and formatting.
        region (Tuple[int, int, int, int]): Table region as
            (start_row, end_row, start_col, end_col).

    Returns:
        pd.DataFrame: Extracted table data with headers applied.
    """
    start_row, end_row, start_col, end_col = region
    merge_map = _build_merge_map(sheet_format, sheet_values)

    headers = []
    for col in range(start_col, end_col + 1):
        v = sheet_values.cell(row=start_row, column=col).value
        if v is None:
            v = merge_map.get((start_row, col))
        f = sheet_format.cell(row=start_row, column=col).value
        if v is None and isinstance(f, str) and f.startswith("="):
            v = f
        headers.append(str(v).strip() if v is not None else f"Column_{col}")

    data = []
    for row in range(start_row + 1, end_row + 1):
        row_data = []
        for col in range(start_col, end_col + 1):
            v = sheet_values.cell(row=row, column=col).value
            if v is None:
                v = merge_map.get((row, col))
            f = sheet_format.cell(row=row, column=col).value
            if v is None and isinstance(f, str) and f.startswith("="):
                v = f
            row_data.append(v)
        if any(v is not None and str(v).strip() != "" for v in row_data):
            data.append(row_data)

    if data:
        return pd.DataFrame(data, columns=headers).dropna(how="all")
    return pd.DataFrame(columns=headers)


def _excel_col_to_index(col: str) -> int:
    """
    Converts an Excel column letter into a zero-based column index.

    Args:
        col (str): Excel column label, such as A, Z, AA, or AB.

    Returns:
        int: Zero-based column index.
    """
    col = col.upper()
    result = 0
    for char in col:
        result = result * 26 + (ord(char) - ord("A") + 1)
    return result - 1


def _read_df_range(df: pd.DataFrame, excel_range: str) -> pd.DataFrame:
    """
    Reads a specific Excel-style range from a DataFrame.

    The function parses a range such as "A1:D20", converts it into DataFrame
    indexes, clips the range safely within the DataFrame shape, and returns the
    selected block.

    Args:
        df (pd.DataFrame): Raw worksheet DataFrame.
        excel_range (str): Excel-style cell range to read.

    Returns:
        pd.DataFrame: DataFrame slice matching the requested range.

    Raises:
        ValueError: If the provided range is not in a valid Excel range format.
    """
    m = re.match(r"([A-Z]+)(\d+):([A-Z]+)(\d+)", excel_range.upper())
    if not m:
        raise ValueError(f"Invalid range: {excel_range}")
    sc_l, sr, ec_l, er = m.groups()
    sc = _excel_col_to_index(sc_l)
    ec = _excel_col_to_index(ec_l)
    sr = int(sr) - 1
    er = int(er)
    max_r, max_c = df.shape
    sr = max(0, min(sr, max_r - 1))
    er = max(0, min(er, max_r))
    sc = max(0, min(sc, max_c - 1))
    ec = max(0, min(ec, max_c - 1))
    return df.iloc[sr:er, sc: ec + 1]


def load_data(file_path: str) -> pd.DataFrame:
    """
    Loads data from a CSV or Excel workbook.

    For CSV files, the function reads the file directly into a DataFrame. For
    Excel files, it attempts to detect formatted tables first. If no formatted
    tables are found, it falls back to value-based table block detection or
    whole-sheet extraction.

    Args:
        file_path (str): Path to the CSV, XLSX, XLSM, or XLS input file.

    Returns:
        pd.DataFrame: Combined DataFrame containing all extracted table data.

    Raises:
        FileNotFoundError: If the input file path does not exist.
        ValueError: If the file type is unsupported, the CSV is empty, or no data
            can be extracted from the workbook.
    """
    path = Path(file_path)
    if not path.exists():
        raise FileNotFoundError(f"Input file not found: {file_path}")

    suffix = path.suffix.lower()

    if suffix == ".csv":
        df = pd.read_csv(file_path)
        if df.empty:
            raise ValueError("The CSV file contains no data.")
        log.info("CSV loaded: %d rows x %d cols from '%s'", *df.shape, file_path)
        return df

    if suffix not in (".xlsx", ".xlsm", ".xls"):
        raise ValueError(f"Unsupported format: {suffix}. Use .xlsx, .xlsm, .xls, or .csv")

    all_frames: List[pd.DataFrame] = []

    if suffix in (".xlsx", ".xlsm"):
        wb_fmt = load_workbook(file_path, data_only=False)
        wb_val = load_workbook(file_path, data_only=True)

        for sheet_name in wb_fmt.sheetnames:
            sheet_fmt = wb_fmt[sheet_name]
            sheet_val = wb_val[sheet_name]

            log.info("Processing sheet: '%s'", sheet_name)
            regions = _detect_formatted_table_regions(sheet_fmt)

            if regions:
                log.info("  -> %d formatted table(s) found", len(regions))
                for t_idx, region in enumerate(regions):
                    df_table = _extract_formatted_table(sheet_val, sheet_fmt, region)
                    if not df_table.empty:
                        df_table["__sheet__"] = sheet_name
                        df_table["__table__"] = t_idx + 1
                        all_frames.append(df_table)
                        log.info("    Table %d: %d rows x %d cols", t_idx + 1, *df_table.shape)
            else:
                log.info("  -> No formatted tables; using fallback block detection")
                df_raw = pd.read_excel(file_path, sheet_name=sheet_name, engine="openpyxl", header=None)

                if df_raw.empty:
                    log.info("    Sheet is empty, skipping.")
                    continue

                sheet_data = [[cell.value for cell in row] for row in sheet_val.iter_rows()]
                ranges = _detect_tables_in_sheet_values(sheet_data)

                if ranges:
                    for t_idx, rng in enumerate(ranges):
                        try:
                            df_block = _read_df_range(df_raw, rng)
                            df_block.columns = df_block.iloc[0].astype(str)
                            df_block = df_block.iloc[1:].reset_index(drop=True)
                            df_block = df_block.dropna(how="all")
                            if not df_block.empty:
                                df_block["__sheet__"] = sheet_name
                                df_block["__table__"] = t_idx + 1
                                all_frames.append(df_block)
                                log.info("    Fallback table %d (%s): %d rows x %d cols", t_idx + 1, rng, *df_block.shape)
                        except Exception as exc:
                            log.warning("    Could not extract range %s: %s", rng, exc)
                else:
                    df_raw.columns = df_raw.iloc[0].astype(str)
                    df_raw = df_raw.iloc[1:].reset_index(drop=True).dropna(how="all")
                    if not df_raw.empty:
                        df_raw["__sheet__"] = sheet_name
                        df_raw["__table__"] = 1
                        all_frames.append(df_raw)
                        log.info("    Whole-sheet fallback: %d rows x %d cols", *df_raw.shape)

    else:
        xls = pd.ExcelFile(file_path, engine="xlrd")
        for sheet_name in xls.sheet_names:
            log.info("Processing sheet (xls): '%s'", sheet_name)
            df_raw = pd.read_excel(file_path, sheet_name=sheet_name, engine="xlrd", header=None)
            if df_raw.empty:
                continue
            ranges = _detect_tables_in_sheet_values(df_raw.values.tolist())
            if ranges:
                for t_idx, rng in enumerate(ranges):
                    try:
                        df_block = _read_df_range(df_raw, rng)
                        df_block.columns = df_block.iloc[0].astype(str)
                        df_block = df_block.iloc[1:].reset_index(drop=True).dropna(how="all")
                        if not df_block.empty:
                            df_block["__sheet__"] = sheet_name
                            df_block["__table__"] = t_idx + 1
                            all_frames.append(df_block)
                    except Exception as exc:
                        log.warning("    Range %s failed: %s", rng, exc)
            else:
                df_raw.columns = df_raw.iloc[0].astype(str)
                df_raw = df_raw.iloc[1:].reset_index(drop=True).dropna(how="all")
                if not df_raw.empty:
                    df_raw["__sheet__"] = sheet_name
                    df_raw["__table__"] = 1
                    all_frames.append(df_raw)

    if not all_frames:
        raise ValueError("No data could be extracted from the workbook.")

    combined = pd.concat(all_frames, ignore_index=True)
    log.info(
        "load_data complete: %d total rows x %d cols from %d table(s) across sheets",
        combined.shape[0], combined.shape[1], len(all_frames),
    )
    # print(combined)
    return combined


_STYLES_CACHE: Dict[str, Any] = {}


def _build_styles() -> Tuple[Any, Any, Any, Any]:
    """
    Builds and caches ReportLab paragraph styles used in the generated PDF.

    The function creates title, section heading, body, and monospace styles.
    Once built, the styles are stored in a module-level cache so repeated calls
    reuse the same style objects.

    Returns:
        Tuple[Any, Any, Any, Any]: Title, heading, body, and monospace styles.
    """
    global _STYLES_CACHE
    if _STYLES_CACHE:
        return (
            _STYLES_CACHE["title"],
            _STYLES_CACHE["h1"],
            _STYLES_CACHE["body"],
            _STYLES_CACHE["mono"],
        )

    base = getSampleStyleSheet()

    title_style = ParagraphStyle(
        "ExcelReportTitle", parent=base["Title"],
        fontSize=22, leading=26,
        textColor=colors.HexColor("#1a1a2e"),
        spaceAfter=6, alignment=TA_CENTER,
    )
    h1_style = ParagraphStyle(
        "ExcelSectionHeader", parent=base["Heading1"],
        fontSize=13, leading=17,
        textColor=colors.HexColor("#16213e"),
        spaceBefore=12, spaceAfter=5,
    )
    body_style = ParagraphStyle(
        "ExcelBody", parent=base["Normal"],
        fontSize=9, leading=13,
        textColor=colors.HexColor("#333333"),
        spaceAfter=5, alignment=TA_LEFT,
    )
    mono_style = ParagraphStyle(
        "ExcelMono", parent=base["Code"],
        fontSize=7.5, leading=10.5,
        textColor=colors.HexColor("#444444"),
        backColor=colors.HexColor("#f5f5f5"),
        spaceAfter=5,
    )

    _STYLES_CACHE = {"title": title_style, "h1": h1_style, "body": body_style, "mono": mono_style}
    return title_style, h1_style, body_style, mono_style
